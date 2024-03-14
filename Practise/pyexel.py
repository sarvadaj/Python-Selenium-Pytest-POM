import openpyxl
book = openpyxl.load_workbook("C:\\Users\\GS-0853\\Downloads\\pythonExcel.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Sarvada"
print(sheet.cell(row=2, column=2).value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value = "TestCase2":

    for j in range(1, sheet.max_column):
        print(sheet.cell(row=i,column=j).value)