import openpyxl
class TestDataHome:
    test_data = [{"firstname": "Sarvada", "lastname": "Jadhav", "gender": "Female"},
        {"firstname": "Swapnil", "lastname": "Vispute", "gender": "Male"}]


#call this method with class name as declared Static and also self parameter in fun declaration is not required
    @staticmethod

    def get_test_data(test_case_name):

        dict = {}

        book = openpyxl.load_workbook("C:\\Users\\GS-0853\\Downloads\\pythonExcel.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column +1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[dict]





